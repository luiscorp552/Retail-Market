import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime, timedelta
import os
import glob
import re
import openpyxl
from openpyxl import load_workbook
import shutil
from selenium.webdriver.chrome.options import Options

# Carpeta de destino para las descargas (definida al inicio para usarla en la configuración del driver)
download_folder = r"C:\Ventas-Inventario_icompras_OTC-RX\ventas"

# Configuración del driver con opciones de descarga personalizada
chrome_options = Options()
prefs = {
    "download.default_directory": download_folder,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
chrome_options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(options=chrome_options)  # Asegúrate de que chromedriver esté en el PATH
driver.maximize_window()
wait = WebDriverWait(driver, 15)  # espera explícita de hasta 15 segundos

try:
    # 1. Login
    driver.get("https://erp.rt.com/")
    
    # Esperar campo usuario y escribir
    usuario_xpath = "/html/body/app-root/app-login/div/div/div[2]/div/div[2]/form/div[1]/input"
    usuario_input = wait.until(EC.presence_of_element_located((By.XPATH, usuario_xpath)))
    usuario_input.send_keys("luis.cornieles@mail.com")
    
    # Campo contraseña
    pass_xpath = "/html/body/app-root/app-login/div/div/div[2]/div/div[2]/form/div[2]/input"
    pass_input = driver.find_element(By.XPATH, pass_xpath)
    pass_input.send_keys("xxxxx")
    
    # Botón ingresar
    btn_xpath = "/html/body/app-root/app-login/div/div/div[2]/div/div[2]/form/div[4]/button/span"
    btn_login = driver.find_element(By.XPATH, btn_xpath)
    btn_login.click()
    
    # Esperar a que termine de cargar la página posterior al login
    time.sleep(10)
    
    wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-layout/div/div/app-panel-topbar/div/div/div[1]/span/app-current-office-selector/div/div[2]/p-dropdown/div/div[2]/span")))
    
    # 2. Navegar a salesreport (primera vez, luego se hará por cada sucursal)
    driver.get("https://erp.rt.com/som/salesreport")
    wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-layout/div/div/app-panel-topbar/div/div/div[1]/span/app-current-office-selector/div/div[2]/p-dropdown/div/div[2]/span")))
    
    # Hacer click en el desplegable de selector de oficina para obtener el total de sucursales (solo informativo)
    dropdown_toggle_xpath = "/html/body/app-root/app-layout/div/div/app-panel-topbar/div/div/div[1]/span/app-current-office-selector/div/div[2]/p-dropdown/div/div[2]/span"
    dropdown_toggle = wait.until(EC.element_to_be_clickable((By.XPATH, dropdown_toggle_xpath)))
    dropdown_toggle.click()
    ul_xpath = "/html/body/app-root/app-layout/div/div/app-panel-topbar/div/div/div[1]/span/app-current-office-selector/div/div[2]/p-dropdown/div/div[3]/div[2]/ul"
    ul_element = wait.until(EC.presence_of_element_located((By.XPATH, ul_xpath)))
    spans = ul_element.find_elements(By.XPATH, ".//span[contains(@class, 'ng-star-inserted')]")
    sucursales_count = len(spans)
    print(f"Cantidad de sucursales encontradas en el sistema: {sucursales_count}")
    # Cerrar el dropdown (click fuera)
    driver.find_element(By.TAG_NAME, 'body').click()
    time.sleep(1)
    
    # Lista de sucursales objetivo
    lista_sucursales = [
        "Sucursal 1",
        "Sucursal 2",
        "Sucursal 3",
        "Sucursal 4",
        "Sucursal 5"
    ]
    
    # ========== PASOS 6, 7 y 8 AGREGADOS (fechas, se calculan una sola vez) ==========
    hoy = datetime.now().date()
    fecha_inicio = hoy - timedelta(days=90)
    fecha_fin = hoy - timedelta(days=1)
    print(f"Fechas a aplicar para todas las sucursales: inicio {fecha_inicio} | fin {fecha_fin}")
    
    # Función auxiliar para seleccionar fecha en un calendario dado su XPath base
    def seleccionar_fecha(calendar_base_xpath, year, month, day):
        btn_cal = calendar_base_xpath + "/span/button"
        wait.until(EC.element_to_be_clickable((By.XPATH, btn_cal))).click()
        time.sleep(0.5)
        btn_year = calendar_base_xpath + "/span/div/div/div/div[1]/div/button[2]"
        wait.until(EC.element_to_be_clickable((By.XPATH, btn_year))).click()
        time.sleep(0.5)
        year_container = calendar_base_xpath + "/span/div/div[2]"
        wait.until(EC.presence_of_element_located((By.XPATH, year_container)))
        year_spans = driver.find_elements(By.XPATH, year_container + "//span[contains(@class, 'p-yearpicker-year')]")
        for span in year_spans:
            if span.text.strip() == str(year):
                span.click()
                break
        time.sleep(0.5)
        month_span_xpath = calendar_base_xpath + f"/span/div/div[2]/span[{month}]"
        wait.until(EC.element_to_be_clickable((By.XPATH, month_span_xpath))).click()
        time.sleep(0.5)
        tbody_xpath = calendar_base_xpath + "/span/div/div/div/div[2]/table/tbody"
        days = driver.find_elements(By.XPATH, tbody_xpath + "//td[contains(@class, 'ng-star-inserted')]")
        for td in days:
            if td.text.strip() == str(day):
                td.click()
                break
        time.sleep(0.5)
    
    # XPath base de los calendarios
    calendar_start_base = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[11]/p-calendar"
    calendar_end_base = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[12]/p-calendar"
    
    # Otras variables útiles (XPath de elementos reutilizables)
    dropdown_selector_xpath = "/html/body/app-root/app-layout/div/div/app-panel-topbar/div/div/div[1]/span/app-current-office-selector/div/div[2]/p-dropdown/div/div[2]"
    ul_oficinas_xpath = "/html/body/app-root/app-layout/div/div/app-panel-topbar/div/div/div[1]/span/app-current-office-selector/div/div[2]/p-dropdown/div/div[3]/div[2]/ul"
    filtro_dropdown_xpath = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[1]/p-dropdown/div/div[2]/span"
    input_busqueda_xpath = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[1]/p-dropdown/div/div[3]/div[1]/div/input"
    opciones_xpath = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[1]/p-dropdown/div/div[3]/div[2]/ul//li/span"
    btn_buscar_xpath = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[15]/button[1]/span[2]"
    btn_excel_xpath = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[15]/button[3]"
    total_xpath = "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[2]/div[2]/span[5]"
    
    # Funciones auxiliares de manejo de archivos y formatos
    def get_latest_file(folder, extension=".xlsx"):
        files = glob.glob(os.path.join(folder, f"*{extension}"))
        if not files:
            return None
        latest = max(files, key=os.path.getmtime)
        return latest
    
    def wait_for_new_file(folder, previous_files_set, timeout=180):
        start_time = time.time()
        last_size = -1
        stable_count = 0
        while time.time() - start_time < timeout:
            current_xlsx = set(glob.glob(os.path.join(folder, "*.xlsx")))
            crdownloads = glob.glob(os.path.join(folder, "*.crdownload"))
            if crdownloads:
                time.sleep(1)
                continue
            new_files = current_xlsx - previous_files_set
            if new_files:
                candidate = max(new_files, key=os.path.getmtime)
                try:
                    current_size = os.path.getsize(candidate)
                    if current_size == last_size:
                        stable_count += 1
                    else:
                        stable_count = 0
                        last_size = current_size
                    if stable_count >= 2:
                        return candidate
                except OSError:
                    pass
                time.sleep(1)
                continue
            time.sleep(1)
        raise TimeoutError("No se detectó nuevo archivo en la carpeta de descarga")
    
    def formatear_valor(valor_texto):
        valor_limpio = valor_texto.replace(" ", "").replace("$", "").strip()
        if "," in valor_limpio and "." in valor_limpio:
            if valor_limpio.rfind(",") > valor_limpio.rfind("."):
                valor_limpio = valor_limpio.replace(".", "").replace(",", ".")
            else:
                valor_limpio = valor_limpio.replace(",", "")
        elif "," in valor_limpio:
            partes = valor_limpio.split(",")
            if len(partes) == 2:
                valor_limpio = valor_limpio.replace(",", ".")
            else:
                ultima_comma = valor_limpio.rfind(",")
                valor_limpio = valor_limpio[:ultima_comma].replace(",", "") + valor_limpio[ultima_comma:].replace(",", ".")
        elif "." in valor_limpio:
            partes = valor_limpio.split(".")
            if len(partes) != 2:
                ultimo_punto = valor_limpio.rfind(".")
                valor_limpio = valor_limpio[:ultimo_punto].replace(".", "") + valor_limpio[ultimo_punto:]
        return float(valor_limpio)
    
    def sum_total_ventas_usd(excel_path):
        wb = load_workbook(excel_path, data_only=True)
        sheet = wb.active
        target_header = "Total ventas USD"
        header_row = None
        col_idx = None
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    clean_value = " ".join(cell.value.strip().split())
                    if clean_value.lower() == target_header.lower():
                        header_row = cell.row
                        col_idx = cell.column
                        break
            if col_idx:
                break
        if col_idx is None:
            raise ValueError(f"No se encontró la columna '{target_header}' en el archivo Excel")
        total = 0.0
        for row in sheet.iter_rows(min_row=header_row+1, min_col=col_idx, max_col=col_idx, max_row=sheet.max_row, values_only=True):
            val = row[0]
            if val is not None:
                if isinstance(val, (int, float)):
                    total += val
                elif isinstance(val, str):
                    val_clean = val.strip()
                    if val_clean:
                        try:
                            if ',' in val_clean and '.' not in val_clean:
                                val_clean = val_clean.replace(',', '.')
                            num = float(val_clean)
                            total += num
                        except ValueError:
                            try:
                                num = formatear_valor(val_clean)
                                total += num
                            except:
                                pass
        return total
    
    max_attempts = 3
    
    # ========== BUCLE PRINCIPAL POR CADA SUCURSAL ==========
    for sucursal_actual in lista_sucursales:
        print(f"\n=== PROCESANDO SUCURSAL: {sucursal_actual} ===")
        
        # 3. Volver a la página de salesreport (para empezar limpio)
        driver.get("https://erp.rt.com/som/salesreport")
        # Esperar a que cargue correctamente (usamos un elemento característico de salesreport)
        wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-layout/div/div/app-breadcrumb/div/div/div/p-breadcrumb/div/ul/li[5]/a/span")))
        time.sleep(2)
        
        # 4. Seleccionar oficina en el primer dropdown
        dropdown_selector = wait.until(EC.element_to_be_clickable((By.XPATH, dropdown_selector_xpath)))
        dropdown_selector.click()
        time.sleep(1)
        ul_oficinas = wait.until(EC.presence_of_element_located((By.XPATH, ul_oficinas_xpath)))
        items = driver.find_elements(By.XPATH, "//p-dropdownitem/li/span[1]")
        encontrada = False
        for item in items:
            if item.text.strip() == sucursal_actual:
                item.click()
                encontrada = True
                print(f"Sucursal seleccionada: {sucursal_actual}")
                break
        if not encontrada:
            print(f"No se encontró la sucursal {sucursal_actual} en el selector. Se omite.")
            continue
        time.sleep(2)
        
        # Verificar si después de cambiar la sucursal se redirigió a home
        if "home" in driver.current_url:
            print("Detectada redirección a home. Volviendo a salesreport...")
            driver.get("https://erp.rt.com/som/salesreport")
            wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-layout/div/div/app-breadcrumb/div/div/div/p-breadcrumb/div/ul/li[5]/a/span")))
            time.sleep(2)
            # Volver a aplicar el filtro de sucursal en el reporte (segundo dropdown)
            filtro_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, filtro_dropdown_xpath)))
            filtro_dropdown.click()
            input_busqueda = wait.until(EC.presence_of_element_located((By.XPATH, input_busqueda_xpath)))
            input_busqueda.clear()
            input_busqueda.send_keys(sucursal_actual)
            time.sleep(1)
            opciones = wait.until(EC.presence_of_all_elements_located((By.XPATH, opciones_xpath)))
            for opcion in opciones:
                if opcion.text.strip() == sucursal_actual:
                    opcion.click()
                    print(f"Filtro aplicado a: {sucursal_actual}")
                    break
        else:
            # 5. Aplicar filtro de sucursal en el reporte (segundo dropdown)
            filtro_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, filtro_dropdown_xpath)))
            filtro_dropdown.click()
            input_busqueda = wait.until(EC.presence_of_element_located((By.XPATH, input_busqueda_xpath)))
            input_busqueda.clear()
            input_busqueda.send_keys(sucursal_actual)
            time.sleep(1)
            opciones = wait.until(EC.presence_of_all_elements_located((By.XPATH, opciones_xpath)))
            for opcion in opciones:
                if opcion.text.strip() == sucursal_actual:
                    opcion.click()
                    print(f"Filtro aplicado a: {sucursal_actual}")
                    break
        
        # Establecer fechas (inicio y fin)
        seleccionar_fecha(calendar_start_base, fecha_inicio.year, fecha_inicio.month, fecha_inicio.day)
        seleccionar_fecha(calendar_end_base, fecha_fin.year, fecha_fin.month, fecha_fin.day)
        print("Fechas seleccionadas.")
        
        # Seleccionar categorías (div[4] y checkboxes)
        try:
            btn_categorias = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[4]/div/p-button/button")))
            btn_categorias.click()
            time.sleep(1)
            checkbox1 = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/p-treetable/div/div[2]/table/tbody/tr[26]/td/p-treetablecheckbox/div/div[2]")))
            checkbox1.click()
            time.sleep(0.5)
            checkbox2 = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/p-treetable/div/div[2]/table/tbody/tr[27]/td/p-treetablecheckbox/div/div[2]")))
            checkbox2.click()
            time.sleep(0.5)
            print("Categorías seleccionadas correctamente.")
            btn_categorias.click()
        except Exception as e:
            print(f"Error al seleccionar categorías: {e}")
        
        # Bucle de reintentos para esta sucursal
        for intento in range(1, max_attempts + 1):
            print(f"\n--- Intento {intento} para {sucursal_actual} ---")
            
            if intento > 1:
                print("Reintentando: refrescando página y reaplicando filtros...")
                driver.refresh()
                driver.get("https://erp.rt.com/som/salesreport")
                time.sleep(3)
                # Reaplicar todo (sucursal, fechas, categorías)
                dropdown_selector = wait.until(EC.element_to_be_clickable((By.XPATH, dropdown_selector_xpath)))
                dropdown_selector.click()
                time.sleep(1)
                items = driver.find_elements(By.XPATH, "//p-dropdownitem/li/span[1]")
                for item in items:
                    if item.text.strip() == sucursal_actual:
                        item.click()
                        break
                time.sleep(2)
                # Verificar nuevamente si se fue a home
                if "home/home-principal" in driver.current_url:
                    print("Reintento: redirección a home. Volviendo a salesreport...")
                    driver.get("https://erp.r.com/som/salesreport")
                    wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-layout/div/div/app-breadcrumb/div/div/div/p-breadcrumb/div/ul/li[5]/a/span")))
                    time.sleep(2)
                filtro_dropdown = wait.until(EC.element_to_be_clickable((By.XPATH, filtro_dropdown_xpath)))
                filtro_dropdown.click()
                input_busqueda = wait.until(EC.presence_of_element_located((By.XPATH, input_busqueda_xpath)))
                input_busqueda.clear()
                input_busqueda.send_keys(sucursal_actual)
                time.sleep(1)
                opciones = wait.until(EC.presence_of_all_elements_located((By.XPATH, opciones_xpath)))
                for opcion in opciones:
                    if opcion.text.strip() == sucursal_actual:
                        opcion.click()
                        break
                seleccionar_fecha(calendar_start_base, fecha_inicio.year, fecha_inicio.month, fecha_inicio.day)
                seleccionar_fecha(calendar_end_base, fecha_fin.year, fecha_fin.month, fecha_fin.day)
                try:
                    btn_categorias = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[1]/app-salesreport-filter/div/div[4]/div/p-button/button")))
                    btn_categorias.click()
                    time.sleep(1)
                    checkbox1 = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/p-treetable/div/div[2]/table/tbody/tr[26]/td/p-treetablecheckbox/div/div[2]")))
                    checkbox1.click()
                    time.sleep(0.5)
                    checkbox2 = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/p-treetable/div/div[2]/table/tbody/tr[27]/td/p-treetablecheckbox/div/div[2]")))
                    checkbox2.click()
                    time.sleep(0.5)
                    btn_categorias.click()
                except Exception as e:
                    print(f"Error al reaplicar categorías: {e}")
            
            # Hacer click en buscar
            btn_buscar = wait.until(EC.element_to_be_clickable((By.XPATH, btn_buscar_xpath)))
            # Obtener texto inicial del total
            time.sleep(5)
            texto_inicial = driver.find_element(By.XPATH, total_xpath).text.strip()
            print(f"Valor inicial del total: {texto_inicial}")
            
            btn_buscar.click()
            
            # Esperar a que aparezca el botón de detalle (indicador de carga completada)
            WebDriverWait(driver, 180).until(EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-layout/div/div/div[1]/app-salesreport-list/div[2]/p-table/div/div/table/tbody/tr[1]/td[1]/button")))
            
            # Esperar cambio en el total
            def total_cambiado(driver):
                nuevo = driver.find_element(By.XPATH, total_xpath).text.strip()
                return nuevo != texto_inicial
            wait.until(total_cambiado)
            valor_total_texto = driver.find_element(By.XPATH, total_xpath).text.strip()
            print(f"Valor total después de buscar: {valor_total_texto}")
            
            # Guardar lista de archivos antes de la descarga
            archivos_antes = set(glob.glob(os.path.join(download_folder, "*.xlsx")))
            
            # Hacer click en botón Excel
            btn_excel = wait.until(EC.element_to_be_clickable((By.XPATH, btn_excel_xpath)))
            btn_excel.click()
            
            # Esperar nuevo archivo
            try:
                nuevo_archivo = wait_for_new_file(download_folder, archivos_antes, timeout=180)
                print(f"Archivo descargado: {nuevo_archivo}")
            except TimeoutError as e:
                print(e)
                continue  # reintentar
            
            # Formatear valor web
            try:
                valor_web = formatear_valor(valor_total_texto)
                print(f"Valor web formateado: {valor_web}")
            except Exception as e:
                print(f"Error al formatear valor web: {e}")
                continue
            
            # Sumar columna del Excel
            try:
                suma_excel = sum_total_ventas_usd(nuevo_archivo)
                print(f"Suma Excel (Total ventas USD): {suma_excel}")
            except Exception as e:
                print(f"Error al leer Excel: {e}")
                continue
            
            # Comparar
            if abs(valor_web - suma_excel) < 0.01:
                print("¡Los valores coinciden!")
                fecha_str = fecha_fin.strftime("%d-%m-%Y")
                nuevo_nombre = f"{sucursal_actual}+{fecha_str}.xlsx"
                nuevo_path = os.path.join(download_folder, nuevo_nombre)
                if os.path.exists(nuevo_path):
                    os.remove(nuevo_path)
                shutil.move(nuevo_archivo, nuevo_path)
                print(f"Archivo renombrado a: {nuevo_path}")
                break  # salir del bucle de reintentos, pasar a la siguiente sucursal
            else:
                print(f"Los valores NO coinciden. Web: {valor_web}, Excel: {suma_excel}. Reintentando...")
        else:
            print(f"Se agotaron los reintentos para la sucursal {sucursal_actual}. Se pasa a la siguiente.")
    
    print("\n=== PROCESO COMPLETADO PARA TODAS LAS SUCURSALES ===")
    
except TimeoutException as e:
    print("Error de tiempo de espera:", e)
except NoSuchElementException as e:
    print("No se encontró algún elemento:", e)
finally:
    time.sleep(3)
    driver.quit()